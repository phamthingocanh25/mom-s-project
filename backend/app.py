# backend/app.py
import os
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import werkzeug.utils
import gc
import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment,PatternFill
from openpyxl.utils import get_column_letter, rows_from_range, cols_from_range
import math
from copy import deepcopy
import re



# --- IMPORT CÁC MODULE XỬ LÝ ---
from data_processor import *

# --- KHỞI TẠO ỨNG DỤNG FLASK ---
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
CORS(app, resources={r"/api/*": {"origins": "https://phamthingocanh25.github.io"}})
#CORS(app)
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
############### Packing list #########
def _safe_float(value, default=0.0):
    """Chuyển đổi giá trị sang float một cách an toàn."""
    try:
        return float(value) if value not in [None, ""] else default
    except (ValueError, TypeError):
        return default

def _render_single_pallet_unit(item_content, raw_data_map, pallet_counter, pkl_data_list):
    """
    Tạo một dòng trong PKL cho một pallet đơn NGUYÊN (tính toán đơn giản).
    """
    key = str(item_content['product_code']).strip() + '||' + str(item_content['product_name']).strip()
    raw_info = raw_data_map.get(key, {})

    qty_per_box_val = _safe_float(raw_info.get('QtyPerBox'))
    box_per_pallet_val = _safe_float(raw_info.get('BoxPerPallet')) 
    w_pc_kgs = _safe_float(raw_info.get('Wpc_kgs'))

    qty_boxes = math.ceil(box_per_pallet_val * 1.0)
    qty_pcs = qty_boxes * qty_per_box_val
    nw_kgs = qty_pcs * w_pc_kgs
    gw_kgs = math.ceil(nw_kgs + (qty_boxes * 0.4) + 50)
    cbm = 1.0 * 1.15 * 1.15 * 0.8

    row = {
        'Item No.': pallet_counter['item_no'],
        'Pallet': f"No.{pallet_counter['pallet_no']:03d}",
        'Part Name': item_content['product_name'],
        'Part No.': item_content['product_code'],
        "Q'ty (boxes)": qty_boxes,
        "Q'ty (pcs)": qty_pcs,
        'W / pc (kgs)': w_pc_kgs,
        'N.W (kgs)': nw_kgs,
        'G.W (kgs)': gw_kgs,
        'MEAS. (m)': "1.15*1.15*0.8",
        'CBM': cbm,
        "Q'ty/box": qty_per_box_val,
        "Box/Pallet": box_per_pallet_val,
        "Box Spec": raw_info.get('BoxSpec', ''),
    }
    pkl_data_list.append(row)
    pallet_counter['item_no'] += 1
    pallet_counter['pallet_no'] += 1
    return qty_pcs

def _render_combined_pallet_block(item_content, raw_data_map, pallet_counter, pkl_data_list, total_block_ratio, product_allocation_map):
    """
    Tạo một khối dòng trong PKL cho một pallet gộp (lẻ).
    *** LOGIC MỚI: PHÂN BỔ DỰA TRÊN TỔNG SỐ DƯ (Total Remainder-Based Allocation) ***
    - Tính toán Q'ty (pcs) chính xác cho từng mảnh lẻ dựa trên tổng số dư đã được
      tính toán trước cho toàn bộ các container.
    """
    total_nw_group = 0
    total_boxes_group = 0
    items_calculated = []
    EPSILON = 1e-6 # Hằng số để so sánh số thực

    # BƯỚC 1: TÍNH TOÁN THÔNG SỐ CHO TỪNG SẢN PHẨM THÀNH PHẦN
    for item in item_content['items']:
        product_code = item['product_code']
        fractional_part = item['quantity'] # Tỷ lệ pallet của phần lẻ này

        # Lấy thông tin đã tính toán trước từ map
        alloc_data = product_allocation_map.get(product_code)
        if not alloc_data:
            print(f"Warning: No allocation data found for product {product_code}. Calculation might be incorrect.")
            continue

        key_item = str(item['product_code']).strip() + '||' + str(item['product_name']).strip()
        raw_info_item = raw_data_map.get(key_item, {})
        
        qty_per_box_item = _safe_float(raw_info_item.get('QtyPerBox'))
        w_pc_kgs_item = _safe_float(raw_info_item.get('Wpc_kgs'))

        # --- LOGIC PHÂN BỔ DỰA TRÊN TỔNG SỐ DƯ ---
        total_fractional_ratio = alloc_data['total_fractional_ratio']
        fractional_remainder_pcs = alloc_data['fractional_remainder_pcs']

        # Cập nhật tỷ lệ đã xử lý để xác định mảnh lẻ cuối cùng
        alloc_data['processed_fractional_ratio'] += fractional_part
        is_last_fraction = alloc_data['processed_fractional_ratio'] >= (total_fractional_ratio - EPSILON)

        qty_pcs_item = 0
        if is_last_fraction:
            # TÍNH THEO PHẦN CÒN LẠI để đảm bảo tổng chính xác, chống sai số làm tròn
            pcs_da_phan_bo = alloc_data['processed_fractional_pcs']
            qty_pcs_item = fractional_remainder_pcs - pcs_da_phan_bo
            qty_pcs_item = max(0, round(qty_pcs_item)) # Đảm bảo không âm
        else:
            # PHÂN BỔ CÂN XỨNG theo tỷ lệ
            if total_fractional_ratio > 0:
                 proportional_pcs = fractional_remainder_pcs * (fractional_part / total_fractional_ratio)
                 qty_pcs_item = round(proportional_pcs)
            else:
                 # Trường hợp hiếm gặp khi không có tỷ lệ lẻ nào
                 qty_pcs_item = 0

        # Cập nhật tracker số pcs đã xử lý cho mảnh lẻ
        alloc_data['processed_fractional_pcs'] += qty_pcs_item
        
        # Tính toán các thông số còn lại dựa trên qty_pcs_item
        qty_boxes_item = math.ceil(qty_pcs_item / qty_per_box_item) if qty_per_box_item > 0 else 0
        nw_kgs_item = qty_pcs_item * w_pc_kgs_item

        items_calculated.append({
            'product_name': item['product_name'],
            'product_code': item['product_code'],
            'qty_boxes': qty_boxes_item,
            'qty_pcs': qty_pcs_item,
            'w_pc_kgs': w_pc_kgs_item,
            'nw_kgs': nw_kgs_item,
            'raw_info': raw_info_item
        })

        total_nw_group += nw_kgs_item
        total_boxes_group += qty_boxes_item

    # BƯỚC 2 & 3: TÍNH TỔNG KHỐI VÀ GHI DỮ LIỆU (Không thay đổi)
    gw_kgs_group = math.ceil(total_nw_group + (total_boxes_group * 0.4) + 50)
    cbm_group = total_block_ratio * 1.15 * 1.15 * 0.8
    is_first_item_in_group = True
    for item_data in items_calculated:
        row = {
            'Item No.': pallet_counter['item_no'] if is_first_item_in_group else '',
            'Pallet': f"No.{pallet_counter['pallet_no']:03d}" if is_first_item_in_group else '',
            'Part Name': item_data['product_name'],
            'Part No.': item_data['product_code'],
            "Q'ty (boxes)": item_data['qty_boxes'],
            "Q'ty (pcs)": item_data['qty_pcs'],
            'W / pc (kgs)': item_data['w_pc_kgs'],
            'N.W (kgs)': item_data['nw_kgs'],
            'G.W (kgs)': gw_kgs_group if is_first_item_in_group else 0,
            'MEAS. (m)': "1.15*1.15*0.8" if is_first_item_in_group else '',
            'CBM': cbm_group if is_first_item_in_group else 0,
            "Q'ty/box": item_data['raw_info'].get('QtyPerBox', ''),
            "Box/Pallet": item_data['raw_info'].get('BoxPerPallet', ''),
            "Box Spec": item_data['raw_info'].get('BoxSpec', ''),
        }
        pkl_data_list.append(row)
        is_first_item_in_group = False
    
    pallet_counter['item_no'] += 1
    pallet_counter['pallet_no'] += 1


def write_packing_list_to_sheet(ws, data_df, container_id_num,cumulative_pallet_count,cumulative_pcs, cumulative_nw, cumulative_gw):
    """
    Ghi dữ liệu Packing List đã được định dạng vào một worksheet của openpyxl.
    Hàm này được viết lại để khớp với yêu cầu chi tiết về cột và định dạng.
    """
    # --- 1. ĐỊNH NGHĨA CÁC STYLE SỬ DỤNG CHUNG ---
    font_main_title = Font(name='Arial', size=48, bold=True)
    font_box_header = Font(name='Arial', size=11, bold=True, underline='single')
    font_content = Font(name='Arial', size=11)
    font_content_bold = Font(name='Arial', size=11, bold=True)
    font_table_header = Font(name='Arial', size=10, bold=True)
    font_table_content = Font(name='Arial', size=10)
    thin_side = Side(style='thin')
    border_top_only = Border(top=thin_side)
    border_top_only = Border(top=Side(style='thin'))
    font_bold = Font(name='Arial', size=12, bold=True)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right_center = Alignment(horizontal='right', vertical='center', wrap_text=True)
    align_top_right_bold = Alignment(horizontal='right', vertical='center', wrap_text=False)
    font_table_header_bold = Font(name='Arial', size=10, bold=True)
    align_bottom_center = Alignment(horizontal='center', vertical='bottom')

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    def apply_border_to_range(cell_range, border_style):
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border_style

    # --- 2. THIẾT LẬP CƠ BẢN CHO TRANG TÍNH ---
    ws.sheet_view.showGridLines = False
    
    column_widths = {
         'D': 10, 'E': 10, 'F': 35, 'G': 35,
        'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 18, 'N': 12, 'O': 12, 'P': 12, 'Q': 12, 'R':18
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # --- 3. TẠO TIÊU ĐỀ CHÍNH "PACKING LIST" ---
    ws.merge_cells('D1:P1')
    title_cell = ws['D1']
    title_cell.value = "PACKING LIST"
    title_cell.font = font_main_title
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 65

    # --- 4. TẠO CÁC Ô THÔNG TIN (SELLER, BUYER, INVOICE, FROM/TO) ---
    ws.merge_cells('D2:H2')
    ws.merge_cells('D3:H10')
    apply_border_to_range('D2:H10', thin_border)
    seller_header_cell = ws['D2']
    seller_header_cell.value = "SELLER"
    seller_header_cell.font = font_box_header
    seller_header_cell.alignment = Alignment(horizontal='left', vertical='top')
    seller_content_cell = ws['D3']
    seller_content = (
        "MINH QUANG IDS TRADING AND INDUSTRIES JOINT STOCK COMPANY\n"
        "Add: Plot CN 4 , Yen My Industrial zone, \n"
        "Tan Lap Commune ,Yen My District, Hung Yen Province, Vietnam\n"
        "Tel: (+84) 42 211 8360\n"
        "Fax: (+84) 43 965 2536"
    )
    seller_content_cell.value = seller_content
    seller_content_cell.font = font_content
    seller_content_cell.alignment = align_top_left

    ws.merge_cells('D11:H11')
    ws.merge_cells('D12:H18')
    apply_border_to_range('D11:H18', thin_border)
    buyer_header_cell = ws['D11']
    buyer_header_cell.value = "BUYER"
    buyer_header_cell.font = font_box_header
    buyer_header_cell.alignment = Alignment(horizontal='left', vertical='top')
    buyer_content_cell = ws['D12']
    buyer_content = ("  \n" " \n" " \n")
    buyer_content_cell.value = buyer_content
    buyer_content_cell.font = font_content
    buyer_content_cell.alignment = align_top_left
    
    ws.merge_cells('I2:P2')
    ws.merge_cells('I3:P10')
    apply_border_to_range('I2:P10', thin_border)
    invoice_header_cell = ws['I2']
    invoice_header_cell.value = "INVOICE NO. & DATE:"
    invoice_header_cell.font = font_box_header
    invoice_header_cell.alignment = Alignment(horizontal='left', vertical='top')
    invoice_content_cell = ws['I3']
    invoice_content = ("  \n" " \n" " \n")
    invoice_content_cell.value = invoice_content
    invoice_content_cell.font = font_content
    invoice_content_cell.alignment = align_top_left

    apply_border_to_range('I11:P18', thin_border)
    ws.merge_cells('I11:K11')
    ws.merge_cells('L11:P11')
    ws.merge_cells('I12:K18')
    ws.merge_cells('L12:P18')
    ws['I11'].value = "FROM:"
    ws['I11'].font = font_box_header
    ws['I11'].alignment = align_center
    ws['L11'].value = "TO:"
    ws['L11'].font = font_box_header
    ws['L11'].alignment = align_center
    from_cell = ws['I12']
    from_cell.value = "Haiphong, Vietnam"
    from_cell.font = font_content
    from_cell.alignment = align_center

    to_cell = ws['L12']
    to_cell.value = " "
    to_cell.font = font_content
    to_cell.alignment = align_center


    # --- 5. TẠO BẢNG DỮ LIỆU SẢN PHẨM ---
    start_row = 20
    start_col = 4 # Cột D

    headers = [
        "Item No.", "Pallet", "Part Name", "Part No.", "Q'ty\n(boxes)", "Q'ty\n(pcs)",
        "W / pc\n(kgs)", "N.W\n(kgs)", "G.W\n(kgs)", "MEAS.\n(m)", "Q'ty/box",
        "Box/Pallet", "Box Spec"
    ]
    end_col = start_col + len(headers) - 1
    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=i, value=header)
        cell.font = font_table_header
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    current_row = start_row + 1
    
    numeric_cols = ["Q'ty (boxes)", "Q'ty (pcs)", "W / pc (kgs)", "N.W (kgs)", "G.W (kgs)", "CBM", "Pallet Ratio"]
    for col in numeric_cols:
        if col in data_df.columns:
            data_df[col] = pd.to_numeric(data_df[col], errors='coerce').fillna(0)
        else:
            data_df[col] = 0

    for _, row_data in data_df.iterrows():
        # <<-- NEW: Xử lý giá trị cho cột mới trước khi ghi
        total_pallet_ratio_val = row_data.get('Total Pallet Ratio')
        if pd.isna(total_pallet_ratio_val):
            total_pallet_ratio_val = '' # Hiển thị ô trống nếu không có giá trị
        row_values = [
            row_data.get('Item No.', ''), row_data.get('Pallet', ''),
            row_data.get('Part Name', ''), row_data.get('Part No.', ''),
            row_data.get("Q'ty (boxes)", 0), row_data.get("Q'ty (pcs)", 0),
            row_data.get('W / pc (kgs)', 0), row_data.get('N.W (kgs)', 0),
            row_data.get('G.W (kgs)', 0), row_data.get('MEAS. (m)', ''),
            row_data.get("Q'ty/box", ''), row_data.get("Box/Pallet", ''),
            row_data.get('Box Spec', '')
        ]
        
        for i, value in enumerate(row_values, start=start_col):
            cell = ws.cell(row=current_row, column=i, value=value)
            cell.font = font_table_content
            cell.border = thin_border
            
            col_letter = get_column_letter(i)
            if col_letter in ['H', 'I']:
                cell.number_format = '#,##0'
                cell.alignment = align_right_center
            elif col_letter == 'J':
                cell.number_format = '#,##0.0000'
                cell.alignment = align_right_center
            elif col_letter in ['K', 'L']:
                cell.number_format = '#,##0.00'
                cell.alignment = align_right_center
            else:
                cell.alignment = align_center
        current_row += 1
    
    # --- 6. DÒNG TỔNG KẾT (TOTAL) ---
    total_row_num = current_row
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row=total_row_num, column=col_idx).border = thin_border

    total_qty_boxes = data_df["Q'ty (boxes)"].sum()
    total_qty_pcs = data_df["Q'ty (pcs)"].sum()
    total_nw = data_df['N.W (kgs)'].sum()
    total_gw = data_df['G.W (kgs)'].sum()
    
    total_label_cell = ws.cell(row=total_row_num, column=start_col, value=f'TOTAL CONTAINER {container_id_num}')
    ws.merge_cells(start_row=total_row_num, start_column=start_col, end_row=total_row_num, end_column=start_col + 3)
    total_label_cell.font = font_bold
    total_label_cell.alignment = align_center

    totals_data = {
        "Q'ty\n(boxes)": {'value': total_qty_boxes, 'format': '#,##0'},
        "Q'ty\n(pcs)": {'value': total_qty_pcs, 'format': '#,##0'},
        "N.W\n(kgs)": {'value': total_nw, 'format': '#,##0.00'},
        "G.W\n(kgs)": {'value': total_gw, 'format': '#,##0.00'}
    }

    for header_text, data in totals_data.items():
        try:
            col_idx = headers.index(header_text) + start_col
            cell = ws.cell(row=total_row_num, column=col_idx, value=data['value'])
            cell.font = font_bold
            cell.number_format = data['format']
            cell.alignment = align_right_center
        except ValueError:
            print(f"LOGIC ERROR: Header '{header_text}' not found in `headers` list.")
            continue

    for i, col_name in enumerate(headers, start=start_col):
        column_letter = get_column_letter(i)
        if col_name == 'Part Name': ws.column_dimensions[column_letter].width = 35
        elif col_name == 'Part No.': ws.column_dimensions[column_letter].width = 35
        elif 'Spec' in col_name: ws.column_dimensions[column_letter].width = 20
        else: ws.column_dimensions[column_letter].width = 14


    # --- 7. PHẦN CASE MARK ---
    case_mark_row = total_row_num + 3
    ws.cell(row=case_mark_row, column=4, value="CASE MARK").font = font_bold
    details_start_row = case_mark_row + 1
    ws.cell(row=details_start_row,     column=5, value="INVOICE NO.:").font = font_content
    ws.cell(row=details_start_row + 1, column=5, value="").font = font_content
    ws.cell(row=details_start_row + 2, column=5, value="MADE IN VIETNAM").font = font_content
    ws.cell(row=details_start_row , column=8, value="Package:").font = font_content 
    ws.cell(row=details_start_row + 1, column=8, value="Quantity:").font = font_content 
    ws.cell(row=details_start_row + 2, column=8, value="N.W:").font = font_content      
    ws.cell(row=details_start_row + 3, column=8, value="G.W:").font = font_content      
    ws.cell(row=details_start_row , column=9, value=f"{cumulative_pallet_count} pallets").font = font_content
    ws.cell(row=details_start_row + 1, column=9, value=f"{int(cumulative_pcs)} pcs").font = font_content 
    ws.cell(row=details_start_row + 2, column=9, value=f"{cumulative_nw:.2f} kgs").font = font_content     
    ws.cell(row=details_start_row + 3, column=9, value=f"{cumulative_gw:.2f} kgs").font = font_content     

    # --- 8. PHẦN CHỮ KÝ ---
    signature_label_row = details_start_row + 8 
    signature_name_row = signature_label_row + 4
    ws.cell(row=signature_label_row, column=12, value="Signature").font = font_content
    signature_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[signature_name_row].height = 35
    ws.merge_cells(start_row=signature_name_row, start_column=11, end_row=signature_name_row, end_column=13)
    cell_thu = ws.cell(row=signature_name_row, column=11)
    cell_thu.value = "NGUYEN DUC THU\nBusiness Director"
    cell_thu.font = font_content
    cell_thu.alignment = signature_alignment
    for i in range(11, 14):
        ws.cell(row=signature_name_row, column=i).border = border_top_only
    ws.merge_cells(start_row=signature_name_row, start_column=15, end_row=signature_name_row, end_column=17)
    cell_giang = ws.cell(row=signature_name_row, column=15)
    cell_giang.value = "LE MINH GIANG\nGeneral Director"
    cell_giang.font = font_content
    cell_giang.alignment = signature_alignment
    for i in range(15, 18):
        ws.cell(row=signature_name_row, column=i).border = border_top_only

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No selected file"}), 400
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            filename = werkzeug.utils.secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            import pandas as pd
            xls = pd.ExcelFile(filepath)
            return jsonify({"success": True, "filepath": filepath, "sheets": xls.sheet_names})
        except Exception as e:
            return jsonify({"success": False, "error": f"Lỗi khi xử lý file: {str(e)}"}), 500
    return jsonify({"success": False, "error": "Định dạng file không hợp lệ"}), 400


# backend/app.py

def _generate_response_from_containers(containers):
    """
    Chuyển đổi danh sách các đối tượng Container thành định dạng JSON
    mà frontend có thể hiểu được.
    """
    response_data = []
    for container in containers:
        container_dict = {
            "id": container.id,
            "main_company": container.main_company,
            "total_quantity": round(container.total_quantity, 4),
            "total_weight": round(container.total_weight, 2),
            "total_logical_pallets": container.total_logical_pallets,
            "contents": []
        }

        for pallet in sorted(container.pallets, key=lambda p: p.id):
            content_block = {}
            # Xử lý pallet gộp
            if pallet.is_combined and len(pallet.original_pallets) > 1:
                content_block['type'] = 'CombinedPallet'
                content_block['id'] = pallet.id
                content_block['quantity'] = round(pallet.quantity, 4)
                content_block['total_weight'] = round(pallet.total_weight, 2)
                content_block['is_cross_ship'] = pallet.is_cross_ship
                content_block['items'] = []
                for sub_pallet in sorted(pallet.original_pallets, key=lambda p: p.id):
                    content_block['items'].append({
                        "id": sub_pallet.id,
                        "product_code": sub_pallet.product_code,
                        "product_name": sub_pallet.product_name,
                        "company": sub_pallet.company,
                        "quantity": round(sub_pallet.quantity, 4),
                        "total_weight": round(sub_pallet.total_weight, 2),
                        "is_split": sub_pallet.is_split,
                        "split_from_id": sub_pallet.split_from_id
                    })
            # Xử lý pallet đơn (không gộp hoặc gộp từ 1 pallet)
            else:
                # Nếu là pallet gộp nhưng chỉ có 1 item (do bị tách ra), vẫn hiển thị như pallet đơn
                single_item = pallet.original_pallets[0] if pallet.is_combined else pallet
                content_block['type'] = 'SinglePallet'
                content_block['id'] = single_item.id
                content_block['product_code'] = single_item.product_code
                content_block['product_name'] = single_item.product_name
                content_block['company'] = single_item.company
                content_block['quantity'] = round(single_item.quantity, 4)
                content_block['total_weight'] = round(single_item.total_weight, 2)
                content_block['is_cross_ship'] = pallet.is_cross_ship # Lấy trạng thái cross-ship từ pallet cha
                content_block['is_split'] = single_item.is_split
                content_block['split_from_id'] = single_item.split_from_id

            container_dict['contents'].append(content_block)
        response_data.append(container_dict)

    return {"success": True, "data": response_data}


@app.route('/api/process', methods=['POST'])
def process_data():
    """
    API endpoint để xử lý và tối ưu hóa việc xếp pallet vào container.
    Hàm này sử dụng pipeline xử lý đầy đủ từ file test_p2.py.
    """
    try:
        data = request.get_json()
        filepath = data.get('filepath')
        sheet_name = data.get('sheetName')

        if not all([filepath, sheet_name]):
            return jsonify({"success": False, "error": "Thiếu thông tin file hoặc sheet."}), 400

        # --- GIAI ĐOẠN 1: TẢI DỮ LIỆU ---
        all_pallets, error = load_and_prepare_pallets(filepath, sheet_name)
        if error:
            return jsonify({"success": False, "error": error}), 400
        if not all_pallets:
            return jsonify({"success": False, "error": "Không có dữ liệu pallet hợp lệ để xử lý."}), 400

        # --- GIAI ĐOẠN 2: THỰC THI PIPELINE TỐI ƯU HÓA (LOGIC TỪ TEST_P2.PY) ---
        
        # BƯỚC 2: TÁCH TOÀN BỘ PALLET THÀNH PHẦN NGUYÊN VÀ LẺ
        integer_pallets, fractional_pallets = split_integer_fractional_pallets(all_pallets)

        # BƯỚC 3: XỬ LÝ PALLET NGUYÊN QUÁ KHỔ
        oversized_containers, regular_sized_integer_pallets, container_id_counter = handle_all_oversized_pallets(
            all_pallets=integer_pallets,
            start_container_id=1
        )

        # BƯỚC 4: XẾP CÁC PALLET NGUYÊN CÒN LẠI
        final_containers = list(oversized_containers)
        final_containers, unplaced_integer_pallets, container_id_counter = pack_integer_pallets(
            regular_sized_integer_pallets,
            final_containers,
            container_id_counter
        )

        # BƯỚC 5: GỘP VÀ XẾP PALLET LẺ
        combined_pallets, uncombined_pallets = combine_fractional_pallets(fractional_pallets)
        pallets_to_pack_fractional = combined_pallets + uncombined_pallets
        unplaced_fractional_pallets = pack_fractional_pallets(pallets_to_pack_fractional, final_containers)

        # BƯỚC 6: VÒNG LẶP XỬ LÝ TOÀN BỘ PALLET CHỜ
        loop_counter = 0
        while unplaced_integer_pallets or unplaced_fractional_pallets:
            loop_counter += 1
            if loop_counter > 20: # Giới hạn an toàn để tránh vòng lặp vô hạn
                print("Warning: Loop limit reached. Breaking.")
                break

            pallets_before_iteration = len(unplaced_integer_pallets) + len(unplaced_fractional_pallets)

            # 6.1: Xử lý pallet nguyên chờ
            if unplaced_integer_pallets:
                unplaced_integer_pallets = try_pack_pallets_into_same_company_containers(unplaced_integer_pallets, final_containers)
                if unplaced_integer_pallets:
                    can_cross_ship_all = check_cross_ship_capacity_for_list(unplaced_integer_pallets, final_containers, unplaced_fractional_pallets)
                    if can_cross_ship_all:
                        unplaced_integer_pallets = handle_unplaced_pallets_with_smart_splitting(unplaced_integer_pallets, final_containers, unplaced_fractional_pallets)
                        if unplaced_integer_pallets:
                            final_containers, container_id_counter = handle_remaining_integers_iteratively(unplaced_integer_pallets, final_containers, container_id_counter)
                            unplaced_integer_pallets = []
                    else:
                        unplaced_integer_pallets = attempt_partial_cross_ship(unplaced_integer_pallets, final_containers, unplaced_fractional_pallets)
                        if unplaced_integer_pallets:
                            unplaced_integer_pallets, final_containers, container_id_counter = create_and_pack_one_new_container(
                                unplaced_integer_pallets, final_containers, container_id_counter, unplaced_fractional_pallets
                            )

            # 6.2: Xử lý pallet lẻ/gộp chờ
            if unplaced_fractional_pallets:
                unplaced_fractional_pallets = try_pack_unplaced_fractionals_same_company(unplaced_fractional_pallets, final_containers)
                if unplaced_fractional_pallets:
                    unplaced_fractional_pallets = repack_unplaced_pallets(unplaced_fractional_pallets, final_containers)
                if unplaced_fractional_pallets:
                   final_containers, container_id_counter, unplaced_fractional_pallets = split_and_fit_leftovers(
                           unplaced_fractional_pallets, final_containers, container_id_counter
                   )
                if unplaced_fractional_pallets:
                  unplaced_fractional_pallets, container_id_counter = cross_ship_remaining_pallets(
                      unplaced_pallets=unplaced_fractional_pallets,
                      containers=final_containers,
                      next_container_id=container_id_counter,
                      unplaced_integer_pallets=unplaced_integer_pallets
                 )

            # 6.3: Kiểm tra tiến triển
            pallets_after_iteration = len(unplaced_integer_pallets) + len(unplaced_fractional_pallets)
            if pallets_after_iteration > 0 and pallets_after_iteration == pallets_before_iteration:
                print("Warning: No progress in packing loop. Breaking to avoid infinite loop.")
                # Thêm log về các pallet còn lại để debug
                if unplaced_integer_pallets:
                    print("Unplaced integer pallets:")
                    for p in unplaced_integer_pallets: print(f"  - {p}")
                if unplaced_fractional_pallets:
                    print("Unplaced fractional pallets:")
                    for p in unplaced_fractional_pallets: print(f"  - {p}")
                break

        # --- GIAI ĐOẠN 3 (MỚI): TỐI ƯU HÓA HỢP NHẤT CUỐI CÙNG (PHASE 4) ---
        fully_optimized_containers = phase_4_final_consolidation(final_containers)

        # --- GIAI ĐOẠN 4: HOÀN THIỆN VÀ TRẢ KẾT QUẢ ---
        # Sắp xếp container theo ID số để đảm bảo thứ tự
        fully_optimized_containers.sort(key=lambda c: int(re.search(r'\d+', c.id).group()))
        # Đánh lại ID cho tuần tự
        for i, container in enumerate(fully_optimized_containers, 1):
            container.id = f"Cont_{i}"

        # Chuyển đổi kết quả sang JSON và trả về
        response_dict = _generate_response_from_containers(fully_optimized_containers)

        final_response = {
            "success": response_dict.get("success", True),
            "results": response_dict.get("data", [])
        }
        
        # Thêm thông tin về các pallet chưa được xếp (nếu có) vào response
        # để frontend có thể hiển thị cảnh báo
        unplaced_info = []
        for p in (unplaced_integer_pallets or []):
            unplaced_info.append(f"Pallet nguyên: {p.id} ({p.quantity} qty, {p.total_weight} wgt)")
        for p in (unplaced_fractional_pallets or []):
             unplaced_info.append(f"Pallet lẻ/gộp: {p.id} ({p.quantity} qty, {p.total_weight} wgt)")

        if unplaced_info:
            final_response["warning"] = "Không thể xếp hết tất cả pallet. Các pallet còn lại là:"
            final_response["unplaced_pallets"] = unplaced_info

        gc.collect()
        return jsonify(final_response)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Đã xảy ra lỗi hệ thống không mong muốn: {str(e)}"}), 500
    
@app.route('/api/generate_packing_list', methods=['POST'])
def generate_packing_list_endpoint():
    print("\n[BACKEND] Bắt đầu xử lý /api/generate_packing_list với logic PHÂN BỔ MỚI.")
    try:
        data = request.get_json()
        optimized_results = data.get('optimized_results')
        original_filepath = data.get('original_filepath')
        sheet_name = data.get('sheet_name')

        if not all([optimized_results, original_filepath, sheet_name]):
            return jsonify({"success": False, "error": "Thiếu dữ liệu để tạo packing list."}), 400

        raw_data_map, error = load_and_map_raw_data_for_pkl(original_filepath, sheet_name)
        if error:
            return jsonify({"success": False, "error": error}), 500

        # --- LOGIC MỚI: BẮT ĐẦU TÍNH TOÁN PHÂN BỔ ---

        # BƯỚC 1: Tiền xử lý và tính toán toàn cục (Global Pre-computation)
        product_allocation_map = defaultdict(lambda: {
            'total_integer_pallets': 0,
            'total_fractional_ratio': 0.0,
            'total_pcs_from_M': 0,
            'pcs_in_full_pallets': 0,
            'fractional_remainder_pcs': 0,
            'processed_fractional_pcs': 0, # Tracker cho các mảnh lẻ để chống sai số làm tròn
            'processed_fractional_ratio': 0.0 # Tracker để xác định mảnh lẻ cuối cùng
        })

        # Thu thập thông tin tổng hợp (pallet nguyên, tỷ lệ lẻ) cho mỗi mã sản phẩm
        for container_data in optimized_results:
            for content_block in container_data.get('contents', []):
                if content_block['type'] == 'SinglePallet':
                    quantity = _safe_float(content_block.get('quantity', 0))
                    integer_part = math.floor(quantity)
                    fractional_part = quantity - integer_part
                    product_code = content_block['product_code']
                    
                    product_allocation_map[product_code]['total_integer_pallets'] += integer_part
                    if fractional_part > 1e-6:
                        product_allocation_map[product_code]['total_fractional_ratio'] += fractional_part

                elif content_block['type'] == 'CombinedPallet':
                    for item in content_block.get('items', []):
                        product_code = item['product_code']
                        quantity = _safe_float(item.get('quantity', 0))
                        product_allocation_map[product_code]['total_fractional_ratio'] += quantity

        # BƯỚC 2: Tính toán số dư chính xác cho phần lẻ
        for product_code, data in product_allocation_map.items():
            key_item = next((k for k in raw_data_map if k.startswith(str(product_code) + '||')), None)
            if not key_item:
                print(f"Warning: Could not find raw data for product code {product_code} during pre-computation.")
                continue
                
            raw_info_item = raw_data_map.get(key_item, {})
            
            qty_per_box = _safe_float(raw_info_item.get('QtyPerBox'))
            box_per_pallet = _safe_float(raw_info_item.get('BoxPerPallet'))
            total_pcs_from_M = _safe_float(raw_info_item.get('TotalPcsFromM', 0))

            pcs_in_full_pallets = data['total_integer_pallets'] * (qty_per_box * box_per_pallet)
            fractional_remainder_pcs = total_pcs_from_M - pcs_in_full_pallets

            data['total_pcs_from_M'] = total_pcs_from_M
            data['pcs_in_full_pallets'] = pcs_in_full_pallets
            data['fractional_remainder_pcs'] = fractional_remainder_pcs
        
        # --- KẾT THÚC LOGIC TÍNH TOÁN PHÂN BỔ ---


        # --- BƯỚC 3: TẠO DỮ LIỆU PACKING LIST CHO TỪNG CONTAINER (RENDER) ---
        print("[BACKEND] Bắt đầu render Packing List từ kết quả tối ưu hóa...")
        global_pallet_counter = {'item_no': 1, 'pallet_no': 1}
        finalized_dfs = {}
        optimized_results.sort(key=lambda x: int(re.search(r'\d+', x['id']).group()))

        for container_data in optimized_results:
            container_id = container_data['id']
            print(f"  - Đang render cho Container {container_id}...")
            container_rows = []
            sorted_contents = sorted(container_data.get('contents', []), key=lambda x: x.get('id', ''))

            for content_block in sorted_contents:
                block_type = content_block.get('type')

                if block_type == 'SinglePallet':
                    quantity = _safe_float(content_block.get('quantity', 0))
                    integer_part = math.floor(quantity)
                    fractional_part = quantity - integer_part

                    # Render các pallet NGUYÊN
                    if integer_part > 0:
                        for _ in range(int(integer_part)):
                            _render_single_pallet_unit(
                                content_block, raw_data_map, global_pallet_counter, container_rows
                            )

                    # Xử lý phần LẺ (nếu có) bằng cách coi nó như một pallet gộp của 1 item
                    if fractional_part > 1e-6:
                        pseudo_combined_block = {
                            'items': [{
                                'product_code': content_block['product_code'],
                                'product_name': content_block['product_name'],
                                'quantity': fractional_part
                            }]
                        }
                        _render_combined_pallet_block(
                            pseudo_combined_block, raw_data_map, global_pallet_counter, container_rows,
                            fractional_part, product_allocation_map
                        )

                elif block_type == 'CombinedPallet':
                    total_block_ratio = _safe_float(content_block.get('quantity', 0))
                    _render_combined_pallet_block(
                        content_block, raw_data_map, global_pallet_counter, container_rows,
                        total_block_ratio, product_allocation_map
                    )

            if container_rows:
                finalized_dfs[container_id] = pd.DataFrame(container_rows)

        # --- BƯỚC 4: GHI KẾT QUẢ RA FILE EXCEL (Không thay đổi) ---
        print("[BACKEND] Đang tạo file Excel từ dữ liệu đã render...")
        wb = Workbook()
        wb.remove(wb.active)

        sorted_container_ids = sorted(finalized_dfs.keys(), key=lambda x: int(re.search(r'\d+', x).group()))

        case_mark_data_map = {}
        cumulative_pallet_count = 0
        cumulative_pcs = 0.0
        cumulative_nw = 0.0
        cumulative_gw = 0.0

        for container_id in sorted_container_ids:
            df = finalized_dfs.get(container_id)
            if df is None or df.empty:
                continue
            
            for col in ["Q'ty (pcs)", 'N.W (kgs)', 'G.W (kgs)']:
                 if col not in df.columns:
                     df[col] = 0
                 else:
                     df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            current_pallet_count = df.loc[df['Pallet'] != '', 'Pallet'].nunique()

            cumulative_pallet_count += current_pallet_count
            cumulative_pcs += df["Q'ty (pcs)"].sum()
            cumulative_nw += df['N.W (kgs)'].sum()
            cumulative_gw += df['G.W (kgs)'].sum()
            
            case_mark_data_map[container_id] = {
                'cumulative_pallet_count': cumulative_pallet_count,
                'cumulative_pcs': cumulative_pcs,
                'cumulative_nw': cumulative_nw,
                'cumulative_gw': cumulative_gw
            }

        for container_id in sorted_container_ids:
            if container_id not in finalized_dfs: continue
            df_for_pkl = finalized_dfs[container_id]
            container_id_num = ''.join(filter(str.isdigit, container_id))
            ws = wb.create_sheet(title=f"PKL_Cont_{container_id_num}")

            current_case_mark_data = case_mark_data_map[container_id]

            write_packing_list_to_sheet(
                ws, 
                df_for_pkl, 
                container_id_num,
                current_case_mark_data['cumulative_pallet_count'],
                current_case_mark_data['cumulative_pcs'],
                current_case_mark_data['cumulative_nw'],
                current_case_mark_data['cumulative_gw']
            )

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        gc.collect()

        print("[BACKEND] Đang gửi file về cho frontend...")
        return send_file(
            excel_buffer, as_attachment=True,
            download_name=f'PackingList_{sheet_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"\n\n[BACKEND] !!! ĐÃ XẢY RA LỖI KHÔNG MONG MUỐN !!!")
        print(error_details)
        return jsonify({"success": False, "error": f"Lỗi nghiêm trọng ở backend: {str(e)}"}), 500


# --- CHẠY ỨNG DỤNG ---
if __name__ == '__main__':
    from waitress import serve
    print("Starting server with Waitress on http://0.0.0.0:5001")
    serve(app, host='0.0.0.0', port=5001)