import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, rows_from_range, cols_from_range
import numpy as np # Cần để xử lý giá trị rỗng

def create_full_packing_list(data_df, output_filename="Packing_List_Final.xlsx"):
    """
    Tạo một tệp Excel Packing List hoàn chỉnh từ đầu đến cuối,
    kết hợp phần đầu, phần thân (chi tiết hàng hóa) và phần chân (tổng kết, chữ ký).

    Args:
        data_df (pd.DataFrame): DataFrame chứa dữ liệu chi tiết của các kiện hàng.
        output_filename (str): Tên của tệp Excel sẽ được tạo.
    """
    try:
        # ==============================================================================
        # PHẦN 1: KHỞI TẠO WORKBOOK VÀ TẠO PHẦN ĐẦU
        # ==============================================================================
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PKL-FCL"
        ws.sheet_view.showGridLines = False

        # --- ĐỊNH NGHĨA CÁC PHONG CÁCH ---
        font_main_title = Font(name='Arial', size=48, bold=True)
        font_box_header = Font(name='Arial', size=11, bold=True, underline='single')
        font_content = Font(name='Arial', size=11)
        font_bold = Font(name='Arial', size=12, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thick_side = Side(style='medium')
        thin_side = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_top_only = Border(top=thin_side)
        
        # --- ĐỊNH DẠNG BỐ CỤC PHẦN ĐẦU ---
        # *** SỬA ĐỔI: Tăng độ rộng cột để vừa tiêu đề "PACKING LIST" ***
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',"Q"]:
            ws.column_dimensions[col].width = 15

        ws.merge_cells('D1:M1')
        title_cell = ws['D1']
        title_cell.value = "PACKING LIST"
        title_cell.font = font_main_title
        title_cell.alignment = align_center
        ws.row_dimensions[1].height = 65 # Tăng chiều cao hàng để vừa chữ
        
        # *** Hàm helper để kẻ viền ngoài ***
        def apply_outline(cell_range):
            rows = list(rows_from_range(cell_range))
            cols = list(cols_from_range(cell_range))

            top_left_cell = ws[rows[0][0]]
            top_right_cell = ws[rows[0][-1]]
            bottom_left_cell = ws[rows[-1][0]]
            bottom_right_cell = ws[rows[-1][-1]]

            # Viền trên
            for cell_coord in rows[0]:
                ws[cell_coord].border = Border(top=thick_side)
            # Viền dưới
            for cell_coord in rows[-1]:
                ws[cell_coord].border = Border(bottom=thick_side)
            # Viền trái
            for cell_coord in cols[0]:
                ws[cell_coord].border = Border(left=thick_side)
            # Viền phải
            for cell_coord in cols[-1]:
                ws[cell_coord].border = Border(right=thick_side)

            # Sửa góc
            top_left_cell.border = Border(top=thick_side, left=thick_side)
            top_right_cell.border = Border(top=thick_side, right=thick_side)
            bottom_left_cell.border = Border(bottom=thick_side, left=thick_side)
            bottom_right_cell.border = Border(bottom=thick_side, right=thick_side)

        # 2. BỐN Ô THÔNG TIN
        # Ô 1: SELLER (D2:H10)
        apply_outline('D2:H10')
        ws.merge_cells('D3:H10')
        ws['D2'].value = "SELLER"
        ws['D2'].font = font_box_header
        content_seller = ws['D3']
        content_seller.value = ("Tên công ty ABC\n"
                                "123 Đường XYZ, Phường 1, Quận 2\n"
                                "Thành phố Hồ Chí Minh, Việt Nam\n"
                                "TEL: 028-1234-5678\n"
                                "FAX: 028-1234-5679")
        content_seller.alignment = align_top_left
        content_seller.font = font_content

        # Ô 2: INVOICE (I2:M10)
        apply_outline('I2:M10')
        ws.merge_cells('I3:M5')
        ws.merge_cells('I7:M10')
        ws['I2'].value = "INVOICE NO. / DATE"
        ws['I2'].font = font_box_header
        ws['I6'].value = "PAYMENT"
        ws['I6'].font = font_box_header
        ws['I3'].value = "INV-2025-001 / June 27, 2025"
        ws['I3'].alignment = align_top_left
        ws['I3'].font = font_content
        ws['I7'].value = "T/T in advance"
        ws['I7'].alignment = align_top_left
        ws['I7'].font = font_content

        # Ô 3: BUYER (D11:H18)
        apply_outline('D11:H18')
        ws.merge_cells('D12:H18')
        ws['D11'].value = "BUYER"
        ws['D11'].font = font_box_header
        content_buyer = ws['D12']
        content_buyer.value = ("Tên khách hàng DEF\n"
                               "456 Đại lộ JQK, Quận 5\n"
                               "Thành phố New York, Hoa Kỳ\n"
                               "TEL: 212-987-6543")
        content_buyer.alignment = align_top_left
        content_buyer.font = font_content
        
        # Ô 4: FROM/TO (I11:M18)
        apply_outline('I11:M18')
        ws.merge_cells('I12:M13')
        ws.merge_cells('I15:M18')
        ws['I11'].value = " FROM:"
        ws['I11'].font = font_box_header
        ws['I14'].value = "TO:"
        ws['I14'].font = font_box_header
        ws['I12'].value = "Cảng Cát Lái, Việt Nam"
        ws['I12'].alignment = align_top_left
        ws['I12'].font = font_content
        ws['I15'].value = "Cảng Los Angeles, Hoa Kỳ"
        ws['I15'].alignment = align_top_left
        ws['I15'].font = font_content

        # ==============================================================================
        # PHẦN 2: XỬ LÝ DỮ LIỆU VÀ TẠO BẢNG HÀNG HÓA
        # ==============================================================================
        df_cleaned = data_df.replace(r'^\s*$', np.nan, regex=True).infer_objects(copy=False)
        df_cleaned.dropna(how='all', inplace=True) # Xóa các hàng hoàn toàn trống    
        df_new = pd.DataFrame()
        df_new['Item No'] = data_df['Item No.']
        df_new['Pallet'] = data_df['Pallet']
        df_new['Part Name'] = data_df['Part Name']
        df_new['Part No'] = data_df['Part No.']
        df_new["Q'ty/box"] = data_df["Q'ty/box"]
        df_new['Quantity( boxes)'] = data_df["Q'ty (boxes)"]
        df_new['Quantity(Pcs)'] = data_df["Q'ty (pcs)"]
        df_new['W/pc(kgs)'] = data_df['W / pc (kgs)']
        df_new['N.W(kgs)'] = data_df['N.W (kgs)']
        df_new['G.W(kgs)'] = data_df['G.W (kgs)']
        df_new['MEAS.(m)'] = data_df['MEAS. (m)']
        df_new["Q'ty/ box"] = data_df["Q'ty/ box"]
        df_new["Box/Pallet"] = data_df["Box/Pallet"]

        box_spec_cols = ['Box Spec', 'Unnamed: 7']
        data_df[box_spec_cols] = data_df[box_spec_cols].fillna('').astype(str)
        df_new['Box Spec'] = data_df[box_spec_cols].apply(lambda x: ' '.join(x.dropna()), axis=1).str.strip()
        
        df_new[''] = '' 
        df_new[' '] = ''

        df_new['pallet_group'] = df_new['Pallet'].ffill()
        grouped = df_new.groupby('pallet_group', sort=False)
        
        start_row = 19
        start_col = 4

        headers = list(df_new.columns.drop('pallet_group'))
        for c_idx, value in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=c_idx, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row = start_row + 1
        for _, group in grouped:
            group_data = group.drop(columns=['pallet_group'])
            num_rows_in_group = len(group_data)

            for r_idx, row_data in enumerate(group_data.itertuples(index=False), start=current_row):
                for c_idx, value in enumerate(row_data, start=start_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            if num_rows_in_group > 1:
                # *** SỬA ĐỔI: Chỉ gộp cột 'Item No' và 'Pallet' theo yêu cầu ***
                merge_cols = ['Item No', 'Pallet']
                for col_name in merge_cols:
                    if col_name in headers:
                        col_idx = headers.index(col_name) + start_col
                        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + num_rows_in_group - 1, end_column=col_idx)
                        # Căn giữa cho ô đã gộp
                        cell_to_align = ws.cell(current_row, col_idx)
                        cell_to_align.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += num_rows_in_group

        total_row_num = current_row
        
        def safe_to_numeric(series):
            return pd.to_numeric(series.astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)

        total_qty_boxes = safe_to_numeric(df_new['Quantity( boxes)']).sum()
        total_qty_pcs = safe_to_numeric(df_new['Quantity(Pcs)']).sum()
        total_nw = safe_to_numeric(df_new['N.W(kgs)']).sum()
        total_gw = safe_to_numeric(df_new['G.W(kgs)']).sum()
        
        total_label_cell = ws.cell(row=total_row_num, column=start_col, value='Total container (1)')
        ws.merge_cells(start_row=total_row_num, start_column=start_col, end_row=total_row_num, end_column=start_col + 4)
        total_label_cell.font = font_bold
        total_label_cell.alignment = align_center

        totals_dict = {'boxes': total_qty_boxes, 'pcs': total_qty_pcs, 'nw': total_nw, 'gw': total_gw}
        ws.cell(row=total_row_num, column=headers.index('Quantity( boxes)') + start_col, value=total_qty_boxes).font = font_bold
        ws.cell(row=total_row_num, column=headers.index('Quantity(Pcs)') + start_col, value=total_qty_pcs).font = font_bold
        ws.cell(row=total_row_num, column=headers.index('N.W(kgs)') + start_col, value=total_nw).font = font_bold
        ws.cell(row=total_row_num, column=headers.index('G.W(kgs)') + start_col, value=total_gw).font = font_bold

        for row in ws.iter_rows(min_row=start_row, max_row=total_row_num, min_col=start_col, max_col=start_col + len(headers) - 1):
            for cell in row:
                cell.border = thin_border
                if not cell.alignment.vertical:
                    cell.alignment = Alignment(vertical='center', horizontal='center')

        for i, col_name in enumerate(headers, start=start_col):
            column_letter = get_column_letter(i)
            if col_name == 'Part Name': ws.column_dimensions[column_letter].width = 35
            elif col_name == 'Part No': ws.column_dimensions[column_letter].width = 25
            elif 'Spec' in col_name: ws.column_dimensions[column_letter].width = 20
            else: ws.column_dimensions[column_letter].width = 14
        
        # ==============================================================================
        # PHẦN 3: THÊM TỔNG KẾT VÀ CHỮ KÝ (ĐÃ SỬA LỖI)
        # ==============================================================================
        
        case_mark_row = total_row_num + 3
        ws.cell(row=case_mark_row, column=4, value="CASE MARK").font = font_bold

        details_start_row = case_mark_row + 1
        ws.cell(row=details_start_row,     column=5, value="INVOICE NO.:").font = font_content
        ws.cell(row=details_start_row + 1, column=5, value="BRIGHT MANUFACTURING CO., LTD.").font = font_content
        ws.cell(row=details_start_row + 2, column=5, value="MADE IN VIETNAM").font = font_content
        ws.cell(row=details_start_row,     column=8, value="Package:").font = font_content
        ws.cell(row=details_start_row + 1, column=8, value="Quantity:").font = font_content
        ws.cell(row=details_start_row + 2, column=8, value="N.W:").font = font_content
        ws.cell(row=details_start_row + 3, column=8, value="G.W:").font = font_content

        ws.cell(row=details_start_row,     column=9, value=f"{int(totals_dict['boxes'])} BOXES").font = font_content
        ws.cell(row=details_start_row + 1, column=9, value=f"{int(totals_dict['pcs'])} PCS").font = font_content
        ws.cell(row=details_start_row + 2, column=9, value=f"{totals_dict['nw']:.2f} KGS").font = font_content
        ws.cell(row=details_start_row + 3, column=9, value=f"{totals_dict['gw']:.2f} KGS").font = font_content

        # --- Phần chữ ký (KHỐI MÃ ĐÃ SỬA LỖI HOÀN CHỈNH) ---
        # Tạo khoảng trống lớn hơn để tránh xung đột bố cục
        signature_label_row = details_start_row + 8 
        signature_name_row = signature_label_row + 4 # Dòng cho tên và chức vụ

        # 1. Thêm nhãn "Signature"
        ws.cell(row=signature_label_row, column=11, value="Signature:").font = font_content

        # 2. Định dạng chung cho chữ ký
        signature_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_top_only = Border(top=Side(style='thin'))

        # 3. Tăng chiều cao của hàng chứa tên và chức vụ
        # Đây là bước quan trọng nhất để hiển thị đủ 2 dòng văn bản
        ws.row_dimensions[signature_name_row].height = 35

        # --- Chữ ký 1 (Gộp cột J, K, L) ---
        ws.merge_cells(start_row=signature_name_row, start_column=10, end_row=signature_name_row, end_column=12) 
        cell_thu = ws.cell(row=signature_name_row, column=10)
        cell_thu.value = "NGUYEN DUC THU\nBusiness Director"
        cell_thu.font = font_content
        cell_thu.alignment = signature_alignment
        
        # Kẻ dòng trên cho cả 3 ô đã gộp
        for i in range(10, 13):
            ws.cell(row=signature_name_row, column=i).border = border_top_only

        # --- Chữ ký 2 (Gộp cột O, P, Q) ---
        ws.merge_cells(start_row=signature_name_row, start_column=15, end_row=signature_name_row, end_column=17) 
        cell_giang = ws.cell(row=signature_name_row, column=15)
        cell_giang.value = "LE MINH GIANG\nGeneral Director"
        cell_giang.font = font_content
        cell_giang.alignment = signature_alignment

        # Kẻ dòng trên cho cả 3 ô đã gộp
        for i in range(15, 18):
            ws.cell(row=signature_name_row, column=i).border = border_top_only
        
        
        # ==============================================================================
        # BƯỚC CUỐI: LƯU TỆP EXCEL
        # ==============================================================================
        wb.save(output_filename)
        print(f"✅ Tệp Packing List '{output_filename}' đã được tạo thành công.")

    except FileNotFoundError:
        print(f"❌ Lỗi: Không tìm thấy tệp đầu vào.")
    except KeyError as e:
        print(f"❌ Lỗi: Cột {e} không tồn tại trong DataFrame đầu vào. Vui lòng kiểm tra lại.")
    except Exception as e:
        print(f"❌ Đã xảy ra lỗi không mong muốn: {e}")

# ==============================================================================
# PHẦN CHẠY THỬ NGHIỆM
# ==============================================================================
if __name__ == '__main__':
    # Dữ liệu mẫu giống như trong file gốc
    data = {
        'Item No.': [1, '', '', 2, 3, ''],
        'Pallet': ['PL.No1', '', '', 'PL.No2', 'PL.No3', ''],
        'Part Name': ['Vỏ nhựa loại A', 'Vỏ nhựa loại B', 'Vỏ nhựa loại C', 'Bản lề kim loại', 'Màn hình LCD 7"', ''],
        'Part No.': ['VN-A001', 'VN-B002', 'VN-C003', 'BL-K015', 'LCD-07-S', ''],
        "Q'ty (boxes)": [10, '', '', 8, 15, ''],
        "Q'ty (pcs)": [1000, '', '', 800, 300, ''],
        'W / pc (kgs)': [0.1, 0.12, 0.11, 0.25, 0.5, ''],
        'N.W (kgs)': [100, '', '', 200, 150, ''],
        'G.W (kgs)': [110, '', '', 220, 165, ''],
        'MEAS. (m)': [1.2, '', '', 1.1, 1.5, ''],
        'Q\'ty/box': [100, 100, 100, 100, 20, ''], # Chú ý: Tên cột Q'ty/box không có khoảng trắng ở giữa
        'Box Spec': ['50x30x20', '50x30x20', '50x30x20', '40x40x30', '60x40x25', ''],
        'Unnamed: 7': ['cm', 'cm', 'cm', 'cm', 'cm', ''],
        # --- THÊM DỮ LIỆU CHO 2 CỘT MỚI ---
        'Q\'ty/ box': [100, 100, 100, 100, 20, ''], # Dữ liệu mẫu cho Q'ty/ box
        'Box/Pallet': [10, 10, 10, 8, 15, '']     # Dữ liệu mẫu cho Box/Pallet
    }
    
    sample_df = pd.DataFrame(data)
    # LƯU Ý QUAN TRỌNG VỀ TÊN CỘT:
    # Trong code của bạn là `data_df["Q'ty/ box"]` (có khoảng trắng)
    # Nhưng trong dữ liệu mẫu gốc là `data["Q'ty/box"]` (không có khoảng trắng)
    # Hãy đảm bảo tên cột nhất quán. Đoạn code dưới đây giả định tên cột không có khoảng trắng.
    create_full_packing_list(sample_df, "My_Company_Packing_List_Updated.xlsx")